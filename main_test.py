from tkinter import *
from datetime import date
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.backend_bases import MouseEvent
import numpy as np
import datetime
import seaborn as sns
import pandas as pd
import pathlib
from openpyxl import Workbook,load_workbook
import openpyxl 
import matplotlib
import mysql.connector
from sklearn.preprocessing import StandardScaler

matplotlib.use("TkAgg")

from matplotlib.figure import Figure
import matplotlib.pyplot as plt

from backend import *
from MySQL import *


background="#f0ddd5"
framebg="#62a7ff"
framefg="#fefbfb"

root=Tk()
root.title("CardioSmart 360")
root.geometry("1450x730+40+30")
root.resizable(False,False)
root.config(bg=background)

global mode_state
global prediction
global prediction1
mode_state = tk.IntVar(value=1)

file=pathlib.Path("Heart_data.xlsx")
if file.exists():
    pass
else:
    # parameters = ["age", "sex", "cp", "trestbps", "chol", "fbs", "restecg", "thalach", "exang", "oldpeak", "slope", "ca", "thal", "smoker"]
    file=Workbook()
    sheet=file.active
    sheet["A1"] = "Registration No."
    sheet["B1"] = "Name"
    sheet["C1"] = "Age"
    sheet["D1"] = "Sex"
    sheet["E1"] = "cp"
    sheet["F1"] = "trestbps"
    sheet["G1"] = "chol"
    sheet["H1"] = "fbs"
    sheet["I1"] = "restecg"
    sheet["J1"] = "thalach"
    sheet["K1"] = "exang"
    sheet["L1"] = "oldpeak"
    sheet["M1"] = "slope"
    sheet["N1"] = "ca"
    sheet["O1"] = "thal"
    sheet["P1"] = "smoker"
    # sheet["Q1"] = "Result"

    file.save("Heart_data.xlsx")

#######################Registration NO.#######################
# Automatic registration number entry system
def update_registration_no():
    file=openpyxl.load_workbook("Heart_data.xlsx")
    sheet=file.active
    row=sheet.max_row

    try:
        max_row_value = sheet.cell(row=row, column=1).value
        ActualRegistration.set(max_row_value + 1)
    except:
        ActualRegistration.set("1")
    

####### Analysis <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

def analysis():
    # Define or ensure global variables are defined
    global root, model, Name, DOB, prediction
    mode_state=tk.IntVar(value=1)

    name = Name.get()
    D1 = Date.get()
    today = datetime.date.today()

    # Ensure DOB is in proper format (e.g., datetime)
    A = today.year - DOB.get()  # Assuming DOB is a date object

    try:
        B = selection()
    except:
        messagebox.showerror("missing", "Please select gender!!")
        return

    try:
        F = selection2()
    except:
        messagebox.showerror("missing", "Please select fbs!!")
        return

    try:
        I = selection3()
    except:
        messagebox.showerror("missing", "Please select exang!!")
        return

    try:
        C = int(cp_combobox.get())
    except:
        messagebox.showerror("missing", "Please select cp!!")
        return

    try:
        G = int(restecg_combobox.get())
    except:
        messagebox.showerror("missing", "Please select restecg!!")
        return

    try:
        K = int(slope_combobox.get())
    except:
        messagebox.showerror("missing", "Please select slope!!")
        return

    try:
        L = int(ca_combobox.get())
    except:
        messagebox.showerror("missing", "Please select ca!!")
        return

    try:
        M = int(thal_combobox.get())
    except:
        messagebox.showerror("missing", "Please select thal!!")
        return

    try:
        D = int(trestbps.get())
        E = int(chol.get())
        H = int(thalach.get())
        J = int(oldpeak.get())
    except:
        messagebox.showerror("missing data", "Few missing data entry!!")
        return

    # Assign smoker status based on the toggle button
    N = 1 if choice == "smoking" else 0

    # Debugging info
    print("A is age:", A)
    print("B is gender :", B)
    print("C is cp:", C)
    print("D is trestbps:", D)
    print("E is chol", E)
    print("F is fbs:", F)
    print("G is restcg:", G)
    print("H is thalach:", H)
    print("I is Exang:", I)
    print("J is oldpeak:", J)
    print("K is slope:", K)
    print("L is ca:", L)
    print("M is thal:", M)
    print("N is smoker:", N)

    # Create a whitebox area (frame) for displaying graphs
    frame = tk.Frame(root, bg="white", width=780, height=600)
    frame.place(x=600, y=223)

    # Initialize the graph index
    graph_index = 0

    def update_graph(graph_index):
        f = Figure(figsize=(4.6, 4.6), dpi=100)
        # Define parameters and their values
        parameters = ["age", "sex", "cp", "trestbps", "chol", "fbs", "restecg", "thalach", "exang", "oldpeak", "slope", "ca", "thal", "smoker"]
        values = [A, B, C, D, E, F, G, H, I, J, K, L, M, N]

        # data = pd.read_csv('heart.csv')  # Load dataset 
        data=pd.read_csv('heart_with_smoker_updated.csv')
        
        # Select only the relevant columns (the parameters)
        data = data[parameters]

        # Define the graphs based on index
        if graph_index == 0:
            # Line plot for all parameters
            a = f.add_subplot(111)
            a.plot(parameters, values, marker='o', linestyle='-', linewidth=1, label='All Parameters')
            a.set_title('Graph 1: Line Plot for All Parameters')
            a.set_ylabel('Value')
            a.set_xlabel('Parameters')
            a.tick_params(axis='x', rotation=45)
            a.grid(True)
            a.legend()
        elif graph_index == 1:
            # Radar chart
            num_vars = len(parameters)
            angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
            values_extended = values + values[:1]
            angles_extended = angles + angles[:1]

            a = f.add_subplot(111, polar=True)
            a.fill(angles_extended, values_extended, color='skyblue', alpha=0.25)
            a.plot(angles_extended, values_extended, color='skyblue', linewidth=2)
            a.set_xticks(angles)
            a.set_xticklabels(parameters)
            a.set_title('Graph 2: Radar Chart for Parameters', size=10, color='black', y=1.1)
            f.subplots_adjust(top=0.85)  # Adjust to fit the title properly within the view
        elif graph_index == 2:
            # Heat Map 
            corr = data.corr()

            a = f.add_subplot(111)
            heatmap = sns.heatmap(corr, annot=False, cmap='coolwarm', ax=a, cbar=True,linewidths=0.1,linecolor="white")
            a.set_title('Graph 3: Correlation Heatmap')
            a.set_xticklabels(a.get_xticklabels(), rotation=45, ha='right')

            # Create a hover effect
            def on_hover(event: MouseEvent):
                """Event handler for mouse hover over the heatmap."""
                if event.inaxes == a:  # Ensure the event is inside the heatmap axes
                    # Get the current mouse position in the data space
                    x, y = int(event.xdata), int(event.ydata)

                    # Ensure the coordinates are within the valid range
                    if 0 <= x < len(parameters) and 0 <= y < len(parameters):
                        # Get the value at the hovered location
                        value = corr.iloc[y, x]

                        # Update the title with the hover information
                        a.set_title(f"Hover: {parameters[y]} vs {parameters[x]} = {value:.2f}")
                        f.canvas.draw_idle()

            # Connect the hover event
            f.canvas.mpl_connect('motion_notify_event', on_hover)

            # Add zoom handler to display full annotations when zoomed in
            def on_zoom(event):
                """Event handler for zoom."""
                if event.name == 'scroll_event':
                    if event.step > 0:  # Zooming in
                        heatmap.set(annot=True)  # Show all values
                    else:  # Zooming out
                        heatmap.set(annot=False)  # Hide all values again
                    f.canvas.draw_idle()

            # Connect the zoom event
            f.canvas.mpl_connect('scroll_event', on_zoom)
        elif graph_index == 3:
            # Box plot for all parameters
            a = f.add_subplot(111)
            a.boxplot([[x] for x in values],notch=True, tick_labels=parameters)
            a.set_title('Graph 4: Box Plot for All Parameters')
            a.set_ylabel('Value')
            a.set_xlabel('Parameters')
            a.tick_params(axis='x', rotation=45)
            a.grid(True)
        
        # Clear the previous canvas content
        for widget in frame.winfo_children():
            widget.destroy()

        # Display the updated figure
        canvas = FigureCanvasTkAgg(f, master=frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # Add a toolbar for zooming
        toolbar = NavigationToolbar2Tk(canvas, frame)
        toolbar.update()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    # Function to switch to the next graph
    def next_graph():
        nonlocal graph_index
        graph_index = (graph_index + 1) % 4  # Cycle through graphs
        update_graph(graph_index)

    # Function to switch to the previous graph
    def prev_graph():
        nonlocal graph_index
        graph_index = (graph_index - 1) % 4  # Cycle through graphs
        update_graph(graph_index)

    # Initial Graph Display
    update_graph(graph_index)

    # Load images for buttons (ensure you have these image files in the same directory or provide the correct path)
    prev_image = tk.PhotoImage(file="Images/previous_button(edited).png")  # Replace with the actual path to your image
    next_image = tk.PhotoImage(file="Images/next_button(edited).png")      # Replace with the actual path to your image

    # Add buttons to switch between graphs and save the graph in the same frame
    button_prev = tk.Button(root, image=prev_image, command=prev_graph, borderwidth=0, highlightthickness=0)
    button_prev.image = prev_image
    # No padding; button will exactly match image size
    button_prev.place(x=602, y=225, width=prev_image.width(), height=prev_image.height())

    button_next = tk.Button(root, image=next_image, command=next_graph, borderwidth=0, highlightthickness=0)
    button_next.image = next_image
    # No padding; button will exactly match image size
    button_next.place(x=978, y=225, width=next_image.width(), height=next_image.height())

    report_frame = tk.Frame(root, bg="white", bd=1, relief=tk.RAISED)

    # Create a Text widget for displaying large amounts of text
    report_text = tk.Text(report_frame, font="Arial 10", bg="white", fg="black", wrap=tk.WORD)
    report_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Create a scrollbar for the Text widget
    scrollbar = tk.Scrollbar(report_frame, command=report_text.yview)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Configure the Text widget to use the scrollbar
    report_text.config(yscrollcommand=scrollbar.set)
    # Dynamically place the report frame when the analysis is clicked
    report_frame.place(x=1170, y=410, width=225, height=260)  # Adjust width and height as needed

    # Input data for prediction
    # input_data = (A, B, C, D, E, F, G, H, I, J, K, L, M, N)
    # input_data_as_numpy_array = np.asarray(input_data)
    # input_data_reshape = input_data_as_numpy_array.reshape(1, -1)
    # prediction=model.predict(input_data_reshape)
    # print(prediction)

    input_data = (A, B, C, D, E, F, G, H, I, J, K, L, M, N)
    input_data_as_numpy_array = np.asarray(input_data)
    input_data_reshape = input_data_as_numpy_array.reshape(1, -1)
    input_data_scaled = scaler.transform(input_data_reshape)
    input_data_selected = selector.transform(input_data_scaled)
    prediction = model.predict(input_data_selected)
    print(prediction)


    # Initialize disease type and risk details
    disease_details = []

    # Age
    if A >= 60:
        disease_details.append("Age-related risk: High (Age ≥ 60 years)")
    elif A >= 45:
        disease_details.append("Age-related risk: Moderate (45-60 years)")
    else:
        disease_details.append("Age-related risk: Low (Age < 45 years)")

    # Gender
    if B == 1:
        disease_details.append("Gender: Male (Higher risk)")
    else:
        disease_details.append("Gender: Female (Lower risk)")

    # Chest Pain Type
    if C == 0:
        disease_details.append("Chest pain: Typical angina (High risk)")
    elif C == 1:
        disease_details.append("Chest pain: Atypical angina (Moderate risk)")
    elif C == 2:
        disease_details.append("Chest pain: Non-anginal (Low risk)")
    else:
        disease_details.append("Chest pain: Asymptomatic")

    # Resting Blood Pressure
    if D >= 140:
        disease_details.append("Resting blood pressure: High (≥ 140 mm Hg)")
    elif D >= 120:
        disease_details.append("Resting blood pressure: Prehypertension (120-139 mm Hg)")
    else:
        disease_details.append("Resting blood pressure: Normal (< 120 mm Hg)")

    # Cholesterol
    if E >= 240:
        disease_details.append("Cholesterol: High (≥ 240 mg/dL)")
    elif E >= 200:
        disease_details.append("Cholesterol: Borderline high (200-239 mg/dL)")
    else:
        disease_details.append("Cholesterol: Normal (< 200 mg/dL)")

    # Fasting Blood Sugar
    if F == 1:
        disease_details.append("Fasting blood sugar: High (> 120 mg/dL) - Risk of diabetes")
    else:
        disease_details.append("Fasting blood sugar: Normal")

    # Rest ECG
    if G == 0:
        disease_details.append("Rest ECG: Normal")
    elif G == 1:
        disease_details.append("Rest ECG: ST-T wave abnormality (Moderate to high risk)")
    else:
        disease_details.append("Rest ECG: Left ventricular hypertrophy (High risk)")

    # Maximum Heart Rate
    if H < 100:
        disease_details.append("Maximum heart rate: Low (< 100 bpm)")
    elif H < 150:
        disease_details.append("Maximum heart rate: Moderate (100-149 bpm)")
    else:
        disease_details.append("Maximum heart rate: High (≥ 150 bpm)")

    # Exercise-induced Angina
    if I == 1:
        disease_details.append("Exercise-induced angina: Present (High risk)")
    else:
        disease_details.append("Exercise-induced angina: Absent")

    # ST Depression (oldpeak)
    if J >= 2.0:
        disease_details.append("Oldpeak: High ST depression (≥ 2.0)")
    elif J >= 1.0:
        disease_details.append("Oldpeak: Moderate ST depression (1.0-1.9)")
    else:
        disease_details.append("Oldpeak: Low ST depression (< 1.0)")

    # Slope
    if K == 2:
        disease_details.append("Slope: Downsloping (High risk)")
    elif K == 1:
        disease_details.append("Slope: Flat (Moderate risk)")
    else:
        disease_details.append("Slope: Upsloping (Low risk)")

    # Number of Major Vessels (ca)
    if L >= 1:
        disease_details.append(f"Number of major vessels: {L} (High risk)")
    else:
        disease_details.append("Number of major vessels: 0 (Low risk)")

    # Thalassemia (thal)
    if M == 3:
        disease_details.append("Thalassemia: Reversible defect (High risk)")
    elif M == 2:
        disease_details.append("Thalassemia: Fixed defect (Moderate risk)")
    else:
        disease_details.append("Thalassemia: Normal")

    # Smoker
    if N == 1:
        disease_details.append("Smoking status: Smoker (Increased risk of heart disease)")
    else:
        disease_details.append("Smoking status: Non-smoker (Lower risk)")

    # Generate detailed report
    detailed_report = "\n".join(disease_details)

    if prediction[0] == 1:
        report_text.delete(1.0, tk.END)
        report_text.insert(tk.END, "You have been classified with heart disease based on the following factors:\n", "bold")
        report_text.insert(tk.END, f"{detailed_report}")
    else:
        report_text.delete(1.0, tk.END)
        report_text.insert(tk.END, "You have been classified as not having heart disease based on the following factors:\n", "bold")
        report_text.insert(tk.END, f"{detailed_report}")

    # Configure the 'bold' tag to make text bold
    report_text.tag_configure("bold", font=("Arial", 10, "bold"))



## Info Window(operated by Info Button)<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

def Info():

    Icon_window=Toplevel(root)
    Icon_window.title("Info")
    Icon_window.geometry("700x600+300+100")

    # icon_image
    icon_image=PhotoImage(file="Images/info.png")
    Icon_window.iconphoto(False,icon_image)

    # Heading
    Label(Icon_window,text="Information related to dataset",font="robot 19 bold").pack(padx=20,pady=20)

    # Info
    Label(Icon_window,text="age - age in years",font="arial 11").place(x=20,y=100)
    Label(Icon_window,text="sex - sex (1 = male; 0 = female)",font="arial 11").place(x=20,y=130)
    Label(Icon_window,text="cp - chest pain type (0 = typical angina; 1 = atypical angina; 2 = non-anginal pain; 3 = asymptomatic)",font="arial 11").place(x=20,y=160)
    Label(Icon_window,text="trestbps - resting blood pressure (in mm Hg on admission to the hospital)",font="arial 11").place(x=20,y=190)
    Label(Icon_window,text="chol - serum cholestoral in mg/dl",font="arial 11").place(x=20,y=220)
    Label(Icon_window,text="fbs - fasting blood sugar > 120 mg/dl (1 = true; 0 = false)",font="arial 11").place(x=20,y=250)
    Label(Icon_window,text="restecg - resting electrocardiographic results (0 = normal; 1 = having ST-T; 2 = hypertrophy)",font="arial 11").place(x=20,y=280)
    Label(Icon_window,text="thalach - maximum heart rate achieved",font="arial 11").place(x=20,y=310)
    Label(Icon_window,text="exang - exercise induced angina (1 = yes; 0 = no)",font="arial 11").place(x=20,y=340)
    Label(Icon_window,text="oldpeak - ST depression induced by exercise relative to rest",font="arial 11").place(x=20,y=370)
    Label(Icon_window,text="slope - the slope of the peak exercise ST segment (0 = upsloping; 1 = flat; 2 = downsloping)",font="arial 11").place(x=20,y=400)
    Label(Icon_window,text="ca - number of major vessels (0-3) colored by flourosopy",font="arial 11").place(x=20,y=430)
    Label(Icon_window,text="thal - 0 = normal; 1 = fixed defect; 2 = reversable defect",font="arial 11").place(x=20,y=460)
    Label(Icon_window,text="smoker - 0 = low to normal risk; 1 = high risk",font="arial 11").place(x=20,y=490)


    Icon_window.mainloop()

### It is used for closing the window
def logout():
    root.destroy()
    import Login


### Save()
def Save():
    global prediction1
    prediction1=prediction
    B2=Name.get()
    C2=Date.get()
    D2=DOB.get()
    today=datetime.date.today()
    E2=today.year-DOB.get()
    
    try:
        F2=selection()
    except:
        messagebox.showerror("Missing data","Please select Gender!!")

    try:
        J2=selection2()
    except:
        messagebox.showerror("Missing data","Please select fbs!!")

    try:
        M2=selection3()
    except:
        messagebox.showerror("Missing data","Please select Exang!!")

    try:
        G2=cp_combobox.get()
    except:
        messagebox.showerror("Missing data","Please select cp!!")

    try:
        K2=restecg_combobox.get()
    except:
        messagebox.showerror("Missing data","Please select restecg!!")
        
    try:
        O2=slope_combobox.get()
    except:
        messagebox.showerror("Missing data","Please select slope!!")

    try:
        P2=ca_combobox.get()
    except:
        messagebox.showerror("Missing data","Please select ca!!")    

    try:
        Q2=thal_combobox.get()
    except:
        messagebox.showerror("Missing data","Please select thal!!")    

    H2=trestbps.get()
    I2=chol.get()
    L2=thalach.get()
    N2=float(oldpeak.get())

    # Retrieve smoker status based on the toggle button
    # R2 = 1 if choice == "smoking" else 0
    R2 = mode_state.get()  # 1 for smoking, 0 for non-smoking

    print(B2)
    print(C2)
    print(D2)
    print(E2)
    print(F2)
    print(G2)
    print(H2)
    print(I2)
    print(J2)
    print(K2)
    print(L2)
    print(M2)
    print(N2)
    print(O2)
    print(P2)
    print(Q2)
    print(R2)

    Save_Data_MySql(B2,C2,int(D2),int(E2),int(F2),int(G2),int(H2),int(I2),int(J2),int(K2),int(L2),int(M2),float(N2),int(O2),int(P2),int(Q2),R2)

    try:
        file = load_workbook("Heart_data.xlsx")
        sheet = file.active
        row = sheet.max_row + 1  # Find the next available row

        # Write data to Excel
        sheet.cell(row=row, column=1, value=row - 1)  # Assuming the first column is registration number
        sheet.cell(row=row, column=2, value=B2)
        sheet.cell(row=row, column=3, value=E2)
        sheet.cell(row=row, column=4, value=F2)
        sheet.cell(row=row, column=5, value=G2)
        sheet.cell(row=row, column=6, value=H2)
        sheet.cell(row=row, column=7, value=I2)
        sheet.cell(row=row, column=8, value=J2)
        sheet.cell(row=row, column=9, value=K2)
        sheet.cell(row=row, column=10, value=L2)
        sheet.cell(row=row, column=11, value=M2)
        sheet.cell(row=row, column=12, value=N2)
        sheet.cell(row=row, column=13, value=O2)
        sheet.cell(row=row, column=14, value=P2)
        sheet.cell(row=row, column=15, value=Q2)
        sheet.cell(row=row, column=16, value=R2)
        # sheet.cell(row=row, column=17, value=int(prediction1[0]))

        # Save workbook after writing
        file.save("Heart_data.xlsx")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save data to Excel: {e}")

    Clear()
    update_registration_no()    # It will recheck registration number and reissue new number

### Update
def Update():
    # Retrieve form values
    B2 = Name.get()  # Name (StringVar)
    C2 = Date.get()  # Date (StringVar)
    D2 = DOB.get()  # DOB (StringVar)
    today = datetime.date.today()
    E2 = today.year - int(DOB.get())  # Age calculation, ensure DOB is an integer

    try:
        F2 = selection()  # Selection returns IntVar, use .get() to get the integer value
    except Exception as e:
        messagebox.showerror("Missing data", f"Please select Gender!! Error: {e}")
        return

    try:
        J2 = selection2() # Selection2 returns IntVar, use .get()
    except Exception as e:
        messagebox.showerror("Missing data", f"Please select FBS!! Error: {e}")
        return

    try:
        M2 = selection3() # Selection3 returns IntVar, use .get()
    except Exception as e:
        messagebox.showerror("Missing data", f"Please select Exang!! Error: {e}")
        return

    # Safely retrieve and validate combo box values
    try:
        G2 = int(cp_combobox.get())  # Ensure G2 is an integer value
    except Exception:
        messagebox.showerror("Missing data", "Please select CP!!")
        return

    try:
        K2 = int(restecg_combobox.get())  # Ensure K2 is an integer value
    except Exception:
        messagebox.showerror("Missing data", "Please select RestECG!!")
        return

    try:
        O2 = int(slope_combobox.get())  # Ensure O2 is an integer value
    except Exception:
        messagebox.showerror("Missing data", "Please select Slope!!")
        return

    try:
        P2 = int(ca_combobox.get())  # Ensure P2 is an integer value
    except Exception:
        messagebox.showerror("Missing data", "Please select CA!!")
        return

    try:
        Q2 = int(thal_combobox.get())  # Ensure Q2 is an integer value
    except Exception:
        messagebox.showerror("Missing data", "Please select Thal!!")
        return

    # Retrieve and validate numeric entries
    try:
        H2 = int(trestbps.get())  # Ensure H2 is an integer value
        I2 = int(chol.get())  # Ensure I2 is an integer value
        L2 = int(thalach.get())  # Ensure L2 is an integer value
        N2 = float(oldpeak.get())  # Ensure N2 is a float value
    except ValueError:
        messagebox.showerror("Missing data", "Please fill in all numeric fields correctly!")
        return

    # Get smoker status
    R2 = mode_state.get()  # 1 for smoking, 0 for non-smoking

    # Debugging prints for SQL and Excel validation
    print(f"Updating with values: {B2, C2, D2, E2, F2, G2, H2, I2, J2, K2, L2, M2, N2, O2, P2, Q2, R2}")

    # Update MySQL database
    try:
        mydb = mysql.connector.connect(host='localhost', user='root', password='Patro202172112', database='Heart_Data')
        mycursor = mydb.cursor()

        # Update query
        update_command = """
        UPDATE data
        SET Name=%s, Date=%s, DOB=%s, age=%s, sex=%s, Cp=%s, trestbps=%s, chol=%s, fbs=%s, restecg=%s, 
            thalach=%s, exang=%s, oldpeak=%s, slope=%s, ca=%s, thal=%s, smoker=%s
        WHERE user=%s
        """
        values = (B2, C2, int(D2), E2, F2, G2, H2, I2, J2, K2, L2, M2, N2, O2, P2, Q2, R2, Registration.get())  # Ensure Registration is retrieved
        mycursor.execute(update_command, values)
        mydb.commit()
        messagebox.showinfo("Update", "User data updated successfully in MySQL!")

    except mysql.connector.Error as err:
        messagebox.showerror("Error", f"Failed to update MySQL database: {err}")
        return
    finally:
        if mydb.is_connected():
            mycursor.close()
            mydb.close()

    # Update Excel
    try:
        file = openpyxl.load_workbook("Heart_data.xlsx")
        sheet = file.active

        # Find the row matching the Registration ID
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            if row[0].value == Registration.get():  # Ensure Registration.get() returns the ID
                reg_number = row[0].row
                break
        else:
            messagebox.showerror("Error", "Registration number not found in Excel!")
            return

        # Update the matching row in Excel
        sheet.cell(row=reg_number, column=2, value=B2)  # Name
        sheet.cell(row=reg_number, column=3, value=E2)  # Age
        sheet.cell(row=reg_number, column=4, value=F2)  # Sex
        sheet.cell(row=reg_number, column=5, value=G2)  # CP
        sheet.cell(row=reg_number, column=6, value=H2)  # Trestbps
        sheet.cell(row=reg_number, column=7, value=I2)  # Chol
        sheet.cell(row=reg_number, column=8, value=J2)  # FBS
        sheet.cell(row=reg_number, column=9, value=K2)  # RestECG
        sheet.cell(row=reg_number, column=10, value=L2)  # Thalach
        sheet.cell(row=reg_number, column=11, value=M2)  # Exang
        sheet.cell(row=reg_number, column=12, value=N2)  # Oldpeak
        sheet.cell(row=reg_number, column=13, value=O2)  # Slope
        sheet.cell(row=reg_number, column=14, value=P2)  # CA
        sheet.cell(row=reg_number, column=15, value=Q2)  # Thal
        sheet.cell(row=reg_number, column=16, value=R2)  # Smoker
        # sheet.cell(row=reg_number, column=17, value=int(prediction1[0]))  # Prediction

        # Save the updated Excel sheet
        file.save("Heart_data.xlsx")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save data to Excel: {e}")
        return

    messagebox.showinfo("Success", "Data updated successfully!")
    Clear()
    update_registration_no()  # Refresh registration number


### Clear
def Clear():
    """Reset all fields in the form to their default states."""
    # Reset mode to default (smoking)
    global mode_state, save_button
    print("save_button exists:", "Yes" if save_button else "No")  # Debugging line
    mode_state.set(1)
    mode.config(image=smoking_icon, activebackground="white")
    # Clear text fields
    Registration.set("")
    Name.set("")
    DOB.set("") 
    trestbps.set("")
    chol.set("")
    thalach.set("")
    oldpeak.set("")
    # Reset combo boxes and radio buttons
    gen.set(0)  # Reset gender radio button (no selection)
    fbs.set(0)  # Reset fasting blood sugar radio button (no selection)
    exang.set(0)  # Reset exercise-induced angina radio button (no selection)
    cp_combobox.set("")  # Reset chest pain combo box
    restecg_combobox.set("")  # Reset resting ECG combo box
    slope_combobox.set("")  # Reset slope combo box
    ca_combobox.set("")  # Reset number of major vessels combo box
    thal_combobox.set("")  # Reset thalassemia combo box


def search():
    text=Registration.get()      # Taking input from entry box
    Clear()   # To clear all the data already available in the entry box and other entry fields
    # save_button.config(state="disabled")     # After clicking on search, the save button will be disabled so that no one can click on it
    # Update_button.config(state="normal")    #  After clicking on search, the Update Details button will be enabled

    file=openpyxl.load_workbook("Heart_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name=row[0]
            # print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]

            # print(reg_no_position)
            # print(reg_number)

    try:        #If input term is not available in the record
        print(str(name))
    except:
        messagebox.showerror("Invalid", "Invalid Registration Number!")
    # reg_no_position is shown as Excel cells like A2,A3,A4,.........,An
    # reg_number is shown as numbers like 2,3,4,5,6,.....,n

    x1 = sheet.cell(row=int(reg_number), column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value
    x13 = sheet.cell(row=int(reg_number), column=13).value
    x14 = sheet.cell(row=int(reg_number), column=14).value
    x15 = sheet.cell(row=int(reg_number), column=15).value
    x16 = sheet.cell(row=int(reg_number), column=16).value
    
    Registration.set(x1)
    Name.set(x2)
    DOB.set(x3)
    gen.set(x4)
    cp_combobox.set(x5)
    trestbps.set(x6)
    chol.set(x7)
    fbs.set(x8)
    restecg_combobox.set(x9)
    thalach.set(x10)
    exang.set(x11)
    oldpeak.set(x12)
    slope_combobox.set(x13)
    ca_combobox.set(x14)
    thal_combobox.set(x15)
    mode_state.set(x16)
    # Update mode_state and button image for smoker/non-smoker
    mode_state.set(x16)
    if mode_state.get() == 1:  # Smoker
        mode.config(image=smoking_icon, activebackground="white")
    else:  # Non-smoker
        mode.config(image=non_smoking_icon, activebackground="white")

# $$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$$

# icon 1
image_icon=PhotoImage(file="Images/icon.png")
root.iconphoto(False,image_icon)

# header section 2
logo=PhotoImage(file="Images/header.png")
myimage=Label(image=logo,bg=background)
myimage.place(x=0,y=0)

# frame 3
Heading_entry=Frame(root,width=800,height=190,bg="#df2d4b")
Heading_entry.place(x=600,y=20)

# Label(Heading_entry,text="Registration No.",font="arial 13",bg="#df2d4b",fg=framefg).place(x=5,y=0)
Label(Heading_entry,text="Date",font="arial 13",bg="#df2d4b",fg=framefg).place(x=430,y=0)

Registration=IntVar()
# Modify the Heading_entry to include a Search Field
Label(Heading_entry, text="Search Registration No.", font="arial 13", bg="#df2d4b", fg=framefg).place(x=5, y=0)

Label(Heading_entry,text="Patient Name",font="arial 13",bg="#df2d4b",fg=framefg).place(x=5,y=90)
Label(Heading_entry,text="Birth Year",font="arial 13",bg="#df2d4b",fg=framefg).place(x=430,y=90)

# Search button
search_image = tk.PhotoImage(file="Images/search4.png")  # Replace with the actual path to your image
search_button = tk.Button(Heading_entry, image=search_image,command=search, borderwidth=0, highlightthickness=0)
search_button.image = search_image
search_button.place(x=362, y=31.5, width=search_image.width(), height=search_image.height())

ActualRegistration = StringVar()

# Add a new Registration Number display field below the Heading_entry
Label(root, text="Registration No.", font="arial 13", bg="#df2d4b", fg=framefg).place(x=450, y=20)
registration_display = Entry(root, textvariable=ActualRegistration, width=10, font="arial 15", bg="#ededed", fg="#222222", bd=0, state="readonly")
registration_display.place(x=453, y=60)

# Ensure the registration number is updated when the app starts
update_registration_no()

Entry_image=PhotoImage(file="Images/Rounded Rectangle 1.png")
Entry_image2=PhotoImage(file="Images/Rounded Rectangle 2.png")
# Label(Heading_entry,image=Entry_image,textvariable=Registration,bg="#df2d4b").place(x=5,y=30)
Label(Heading_entry,image=Entry_image,bg="#df2d4b").place(x=5,y=30)
Label(Heading_entry,image=Entry_image,bg="#df2d4b").place(x=430,y=30)

Label(Heading_entry,image=Entry_image2,bg="#df2d4b").place(x=5,y=120)
Label(Heading_entry,image=Entry_image2,bg="#df2d4b").place(x=430,y=120)

search_entry = Entry(Heading_entry, textvariable=Registration, font="arial 15", bg="#0e5363", fg="white", bd=0)
search_entry.place(x=30, y=45)

Date=StringVar()
today=date.today()
d1=today.strftime("%d/%m/%Y")
date_entry=Entry(Heading_entry,textvariable=Date,width=15,font="arial 15",bg="#0e5363",fg="white",bd=0)
date_entry.place(x=440,y=45)
Date.set(d1)

Name=StringVar()
name_entry=Entry(Heading_entry,textvariable=Name,width=20,font="arial 20",bg="#ededed",fg="#222222",bd=0)
name_entry.place(x=30,y=130)

DOB=IntVar()
dob_entry=Entry(Heading_entry,textvariable=DOB,width=20,font="arial 20",bg="#ededed",fg="#222222",bd=0)
dob_entry.place(x=440,y=130)

############################################################# Body ####################################################################
Detail_entry=Frame(root,width=490,height=260,bg="#dbe0e3")
Detail_entry.place(x=30,y=450)

########################radio button###########################
Label(Detail_entry,text="sex:",font="arial 13",bg=framebg,fg=framefg).place(x=10,y=10)
Label(Detail_entry,text="fbs:",font="arial 13",bg=framebg,fg=framefg).place(x=180,y=10)
Label(Detail_entry,text="exang:",font="arial 13",bg=framebg,fg=framefg).place(x=335 ,y=10)

def selection():
    if gen.get()==1:
        Gender=1
        return(Gender)
        print(Gender)
    elif gen.get()==2:
        Gender=2
        return(Gender)
        print(Gender)
    else:
        print(Gender)

def selection2():
    if fbs.get()==1:
        Fbs=1
        return(Fbs)
        print(Fbs)
    elif fbs.get()==2:
        Fbs=2
        return(Fbs)
        print(Fbs)
    else:
        print(Fbs)

def selection3():
    if exang.get()==1:
        Exang=1
        return(Exang)
        print(Exang)
    elif exang.get()==2:
        Exang=2
        return(Exang)
        print(Exang)
    else:
        print(Exang)


gen=IntVar()
R1=Radiobutton(Detail_entry,text="Male",variable=gen,value=1,command=selection)
R2=Radiobutton(Detail_entry,text="Female",variable=gen,value=2,command=selection)
R1.place(x=43,y=10)
R2.place(x=93,y=10)

fbs=IntVar()
R3=Radiobutton(Detail_entry,text="True",variable=fbs,value=1,command=selection2)
R4=Radiobutton(Detail_entry,text="False",variable=fbs,value=2,command=selection2)
R3.place(x=213,y=10)
R4.place(x=263,y=10)

exang=IntVar()
R5=Radiobutton(Detail_entry,text="Yes",variable=exang,value=1,command=selection3)
R6=Radiobutton(Detail_entry,text="No",variable=exang,value=2,command=selection3)
R5.place(x=387,y=10)
R6.place(x=430,y=10)

############################ComboBox############################
Label(Detail_entry,text="cp:",font="arial 13",bg=framebg,fg=framefg).place(x=10,y=50)
Label(Detail_entry,text="restecg:",font="arial 13",bg=framebg,fg=framefg).place(x=10,y=90)
Label(Detail_entry,text="slope:",font="arial 13",bg=framebg,fg=framefg).place(x=10,y=130)
Label(Detail_entry,text="ca:",font="arial 13",bg=framebg,fg=framefg).place(x=10,y=170)
Label(Detail_entry,text="thal:",font="arial 13",bg=framebg,fg=framefg).place(x=10,y=210)

cp_combobox=Combobox(Detail_entry,values=['0','1','2','3'],font="arial 12",state="r",width=14)
restecg_combobox=Combobox(Detail_entry,values=['0','1','2'],font="arial 12",state="r",width=11)
slope_combobox=Combobox(Detail_entry,values=['0','1','2'],font="arial 12",state="r",width=12)
ca_combobox=Combobox(Detail_entry,values=['0','1','2','3','4'],font="arial 12",state="r",width=14)
thal_combobox=Combobox(Detail_entry,values=['0','1','2','3'],font="arial 12",state="r",width=14)

cp_combobox.place(x=50,y=50)
restecg_combobox.place(x=80,y=90)
slope_combobox.place(x=70,y=130)
ca_combobox.place(x=50,y=170)
thal_combobox.place(x=50,y=210)

###############################Data Entry Box####################################
Label(Detail_entry,text="Smoking:",font="arial 13",width=7,bg="#dbe0e3",fg="black").place(x=240,y=50)
Label(Detail_entry,text="trestbps:",font="arial 13",width=7,bg=framebg,fg=framefg).place(x=240,y=90)
Label(Detail_entry,text="chol:",font="arial 13",width=7,bg=framebg,fg=framefg).place(x=240,y=130)
Label(Detail_entry,text="thalach:",font="arial 13",width=7,bg=framebg,fg=framefg).place(x=240,y=170)
Label(Detail_entry,text="oldpeak:",font="arial 13",width=7,bg=framebg,fg=framefg).place(x=240,y=210)

trestbps=StringVar()
chol=StringVar()
thalach=StringVar()
oldpeak=StringVar()

trestbps_entry=Entry(Detail_entry,textvariable=trestbps,width=10,font="arial 15",bg="#ededed",fg="#222222",bd=0)
chol_entry=Entry(Detail_entry,textvariable=chol,width=10,font="arial 15",bg="#ededed",fg="#222222",bd=0)
thalach_entry=Entry(Detail_entry,textvariable=thalach,width=10,font="arial 15",bg="#ededed",fg="#222222",bd=0)
oldpeak_entry=Entry(Detail_entry,textvariable=oldpeak,width=10,font="arial 15",bg="#ededed",fg="#222222",bd=0)
trestbps_entry.place(x=320,y=90)
chol_entry.place(x=320,y=130)
thalach_entry.place(x=320,y=170)
oldpeak_entry.place(x=320,y=210)

######################################################################################################################

############################### Report ##################################

square_report_image=PhotoImage(file="Images/Report(edited).png")
report_background=Label(image=square_report_image,bg=background)
report_background.place(x=1120,y=340)

######################################################################################################################

############################### Graph ####################################

graph_image=PhotoImage(file="Images/graph.png")
Label(image=graph_image).place(x=600,y=270)
Label(image=graph_image).place(x=840,y=270)
Label(image=graph_image).place(x=600,y=500)
Label(image=graph_image).place(x=840,y=500)

############################### Button ####################################

analysis_button=PhotoImage(file="Images/Analysis.png")
Button(image=analysis_button,bd=0,bg=background,cursor="hand2",command=analysis).place(x=1130,y=255)

######################## Info Button ##########################

info_button=PhotoImage(file="Images/info.png")
Button(image=info_button,bd=0,bg=background,cursor="hand2",command=Info).place(x=10,y=240)

######################## Save Button ##########################
global save_button
save_button=PhotoImage(file="Images/save-file_10057635.png")
Button(image=save_button,bd=0,bg=background,cursor="hand2",command=Save).place(x=1370,y=250)

######################## Clear Button ##########################
clear_button=PhotoImage(file="Images/refresh-arrow_11507214.png")
Button(image=clear_button,bd=0,bg=background,cursor="hand2",command=Clear).place(x=1310,y=250)

######################## Update Button ##########################
update_button=PhotoImage(file="Images/file-update_14682609.png")
Button(image=update_button,bd=0,bg=background,cursor="hand2",command=Update).place(x=1250,y=250)

############################## Smoker and Non Smoker Button ################################

# button_mode=True
choice="smoking"
def changemod():
    global mode_state
    if mode_state.get() == 1:  # If currently smoking
        mode_state.set(0)  # Switch to non-smoking
        mode.config(image=non_smoking_icon, activebackground="white")
        choice="smoking"
    else:  # If currently non-smoking
        mode_state.set(1)  # Switch to smoking
        mode.config(image=smoking_icon, activebackground="white")
        choice="Non-smoking"
    print("Current mode:", "Smoking" if mode_state.get() == 1 else "Non-smoking")


smoking_icon=PhotoImage(file="Images/smoker.png")
non_smoking_icon=PhotoImage(file="Images/non-smoker.png")
mode=Button(root,image=smoking_icon,bg="#dbe0e3",bd=0,cursor="hand2",command=changemod)
mode.place(x=350,y=495)

######################################################################################################################

############################## Log Out Button ################################

logout_icon=PhotoImage(file="Images/logout.png")
logout_button=Button(root,image=logout_icon,bg="#df2d4b",cursor="hand2",bd=0,command=logout)
logout_button.place(x=1390,y=60)

######################################################################################################################

root.mainloop()
